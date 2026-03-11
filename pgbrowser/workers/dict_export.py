"""
Data Dictionary export worker.

Runs in a dedicated QThread and opens its own psycopg2 connection so
it never blocks the main-thread UI.  Emits progress signals so the
caller can drive a progress bar, and emits finished/error when done.

Excel layout (one sheet per table):
  Row 1  — merged title:  "Data Dictionary — schema.table"
  Row 2  — frozen header row (dark blue)
  Row 3+ — one row per column, PK cells highlighted yellow, FK cells blue
"""

from PyQt5.QtCore import QThread, pyqtSignal

try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
except ImportError:
    psycopg2 = None       # type: ignore[assignment]
    RealDictCursor = None  # type: ignore[assignment,misc]

from pgbrowser.db.queries import get_columns_for_export


class DictExportWorker(QThread):
    """Generate a multi-sheet Excel data-dictionary for a set of tables.

    Signals
    -------
    progress(current, total, table_name)
        Emitted for each table as it is processed.
    finished(data: bytes)
        Emitted when generation completes; *data* is the raw .xlsx content.
    error(message: str)
        Emitted if an unrecoverable error occurs.
    """

    progress = pyqtSignal(int, int, str)  # (current_index, total, table_name)
    finished = pyqtSignal(bytes)          # raw .xlsx file contents
    error    = pyqtSignal(str)            # human-readable error message

    def __init__(self, params: dict, schema: str, tables: list):
        super().__init__()
        self.params = params
        self.schema = schema
        self.tables = tables

    # ── Thread entry point ─────────────────────────────────────────────────────

    def run(self) -> None:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            self.error.emit("openpyxl is not installed.\nRun:  pip install openpyxl")
            return

        conn = self._open_connection()
        if conn is None:
            return

        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)     # drop the default empty sheet

            # ── Shared styles ──────────────────────────────────────────────────
            hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
            hdr_fill  = PatternFill("solid", fgColor="1F3864")
            hdr_align = Alignment(horizontal="center", vertical="center",
                                  wrap_text=True)
            thin      = Side(style="thin", color="D9D9D9")
            brd       = Border(left=thin, right=thin, top=thin, bottom=thin)
            pk_fill   = PatternFill("solid", fgColor="FFF2CC")
            fk_fill   = PatternFill("solid", fgColor="DEEAF1")

            HEADERS    = [
                "#", "Column Name", "Data Type", "UDT Name",
                "Max Length", "Precision", "Scale",
                "Nullable", "Default Value", "Constraints", "Description",
            ]
            COL_WIDTHS = [5, 25, 18, 18, 12, 12, 8, 10, 30, 18, 40]

            for idx, table in enumerate(self.tables, 1):
                self.progress.emit(idx, len(self.tables), table)
                rows = get_columns_for_export(conn, self.schema, table)
                ws   = self._create_sheet(wb, table, idx)

                # Title row
                ws.merge_cells("A1:K1")
                title_cell           = ws["A1"]
                title_cell.value     = f"Data Dictionary — {self.schema}.{table}"
                title_cell.font      = Font(name="Calibri", bold=True,
                                            size=13, color="1F3864")
                title_cell.alignment = Alignment(horizontal="left",
                                                  vertical="center")
                ws.row_dimensions[1].height = 22

                # Header row
                for ci, (hdr, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
                    cell            = ws.cell(row=2, column=ci, value=hdr)
                    cell.font       = hdr_font
                    cell.fill       = hdr_fill
                    cell.alignment  = hdr_align
                    cell.border     = brd
                    ws.column_dimensions[cell.column_letter].width = w
                ws.row_dimensions[2].height = 18

                # Data rows
                for ri, row in enumerate(rows, 3):
                    constraints = row["constraints"] or ""
                    vals = [
                        row["ordinal_position"],
                        row["column_name"],
                        row["data_type"],
                        row["udt_name"],
                        row["character_maximum_length"],
                        row["numeric_precision"],
                        row["numeric_scale"],
                        row["is_nullable"],
                        row["column_default"] or "",
                        constraints,
                        row["description"] or "",
                    ]
                    for ci, val in enumerate(vals, 1):
                        cell           = ws.cell(row=ri, column=ci, value=val)
                        cell.border    = brd
                        cell.alignment = Alignment(
                            vertical="center",
                            wrap_text=(ci in (9, 11)),
                        )
                        if ci == 10:
                            if "PK" in constraints:
                                cell.fill = pk_fill
                            elif "FK" in constraints:
                                cell.fill = fk_fill
                    ws.row_dimensions[ri].height = 15

                ws.freeze_panes = "A3"

            import io
            buf = io.BytesIO()
            wb.save(buf)
            self.finished.emit(buf.getvalue())

        except Exception as exc:
            self.error.emit(str(exc))
        finally:
            try:
                conn.close()
            except Exception:
                pass

    # ── Helpers ────────────────────────────────────────────────────────────────

    def _open_connection(self):
        try:
            return psycopg2.connect(
                host=self.params["host"],
                port=int(self.params["port"]),
                dbname=self.params["database"],
                user=self.params["user"],
                password=self.params["password"],
                connect_timeout=10,
                application_name="pgbrowser_export",
            )
        except Exception as exc:
            self.error.emit(f"Connection failed: {exc}")
            return None

    @staticmethod
    def _create_sheet(wb, table: str, idx: int):
        """Return a new worksheet with a safe, unique name (≤ 31 chars)."""
        name       = table[:31]
        used_names = {ws.title for ws in wb.worksheets}
        if name in used_names:
            name = f"{table[:28]}_{idx}"
        return wb.create_sheet(title=name)
