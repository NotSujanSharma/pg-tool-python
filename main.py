#!/usr/bin/env python3
"""
PG Browser — PostgreSQL Schema Explorer
A modern PyQt5 GUI for exploring PostgreSQL databases.
"""

import sys
import json
import os

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QStackedWidget,
    QVBoxLayout, QHBoxLayout, QFormLayout,
    QLabel, QLineEdit, QPushButton, QCheckBox,
    QListWidget, QListWidgetItem,
    QTableWidget, QTableWidgetItem,
    QComboBox, QSplitter, QFrame,
    QHeaderView, QAbstractItemView,
    QMessageBox, QSizePolicy, QTabWidget,
    QProgressBar, QFileDialog,
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QFont, QColor

try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
    HAS_PSYCOPG2 = True
except ImportError:
    HAS_PSYCOPG2 = False

PROFILES_FILE = os.path.expanduser("~/.config/pgbrowser/profiles.json")

# ── Stylesheet ─────────────────────────────────────────────────────────────────
STYLESHEET = """
QWidget {
    background-color: #1e1e2e;
    color: #cdd6f4;
    font-family: "Segoe UI", "Noto Sans", Arial, sans-serif;
    font-size: 13px;
}

/* ── Login card ── */
QFrame#card {
    background-color: #181825;
    border: 1px solid #313244;
    border-radius: 12px;
}

/* ── Inputs ── */
QLineEdit {
    background-color: #313244;
    border: 1px solid #45475a;
    border-radius: 6px;
    padding: 7px 10px;
    color: #cdd6f4;
}
QLineEdit:focus {
    border-color: #89b4fa;
}
QLineEdit:disabled {
    color: #585b70;
}
QLineEdit[echoMode="2"] {
    lineedit-password-character: 9679;
}

QComboBox {
    background-color: #313244;
    border: 1px solid #45475a;
    border-radius: 6px;
    padding: 7px 10px;
    color: #cdd6f4;
    min-width: 120px;
}
QComboBox:focus { border-color: #89b4fa; }
QComboBox::drop-down { border: none; width: 24px; }
QComboBox QAbstractItemView {
    background-color: #313244;
    border: 1px solid #45475a;
    selection-background-color: #45475a;
    color: #cdd6f4;
    outline: none;
}

/* ── Buttons ── */
QPushButton {
    background-color: #89b4fa;
    color: #1e1e2e;
    border: none;
    border-radius: 6px;
    padding: 8px 18px;
    font-weight: 600;
    min-width: 80px;
}
QPushButton:hover  { background-color: #b4befe; }
QPushButton:pressed { background-color: #74c7ec; }
QPushButton:disabled { background-color: #45475a; color: #6c7086; }

QPushButton#secondary {
    background-color: #313244;
    color: #cdd6f4;
}
QPushButton#secondary:hover { background-color: #45475a; }

QPushButton#danger {
    background-color: #f38ba8;
    color: #1e1e2e;
}
QPushButton#danger:hover { background-color: #eba0ac; }

/* ── List ── */
QListWidget {
    background-color: #181825;
    border: none;
    outline: none;
    padding: 4px 2px;
}
QListWidget::item {
    padding: 7px 12px;
    border-radius: 5px;
    margin: 1px 2px;
}
QListWidget::item:hover    { background-color: #313244; }
QListWidget::item:selected {
    background-color: #45475a;
    color: #89b4fa;
    font-weight: 600;
}

/* ── Table ── */
QTableWidget {
    background-color: #181825;
    border: none;
    outline: none;
    gridline-color: #313244;
    alternate-background-color: #1e1e2e;
}
QTableWidget::item { padding: 5px 10px; }
QTableWidget::item:selected {
    background-color: #313244;
    color: #cdd6f4;
}
QHeaderView::section {
    background-color: #1e1e2e;
    color: #89b4fa;
    font-weight: 700;
    padding: 7px 10px;
    border: none;
    border-bottom: 2px solid #313244;
    border-right: 1px solid #2a2a3e;
}
QHeaderView::section:last { border-right: none; }

/* ── Splitter ── */
QSplitter::handle           { background-color: #313244; }
QSplitter::handle:horizontal { width: 1px; }

/* ── Tabs ── */
QTabWidget::pane {
    border: 1px solid #313244;
    border-top: none;
    background: #1e1e2e;
}
QTabBar::tab {
    background: #181825;
    color: #a6adc8;
    padding: 8px 22px;
    border: 1px solid #313244;
    border-bottom: none;
    border-radius: 6px 6px 0 0;
    margin-right: 2px;
}
QTabBar::tab:selected {
    background: #1e1e2e;
    color: #89b4fa;
    font-weight: 600;
}
QTabBar::tab:hover:!selected { background: #282838; }

/* ── Scrollbars ── */
QScrollBar:vertical {
    background: transparent; width: 8px; margin: 0;
}
QScrollBar::handle:vertical {
    background: #45475a; border-radius: 4px; min-height: 30px;
}
QScrollBar::handle:vertical:hover { background: #585b70; }
QScrollBar::add-line:vertical,
QScrollBar::sub-line:vertical { height: 0; }

QScrollBar:horizontal {
    background: transparent; height: 8px; margin: 0;
}
QScrollBar::handle:horizontal {
    background: #45475a; border-radius: 4px; min-width: 30px;
}
QScrollBar::handle:horizontal:hover { background: #585b70; }
QScrollBar::add-line:horizontal,
QScrollBar::sub-line:horizontal { width: 0; }

/* ── Groupbox ── */
QGroupBox {
    border: 1px solid #313244;
    border-radius: 8px;
    margin-top: 14px;
    padding-top: 6px;
}
QGroupBox::title {
    subcontrol-origin: margin;
    subcontrol-position: top left;
    left: 12px; padding: 0 4px;
    color: #89b4fa; font-weight: 700;
}

/* ── Checkbox ── */
QCheckBox { spacing: 6px; }
QCheckBox::indicator {
    width: 16px; height: 16px;
    border-radius: 4px; border: 1px solid #45475a; background: #313244;
}
QCheckBox::indicator:checked { background: #89b4fa; border-color: #89b4fa; }

/* ── Misc ── */
QMessageBox, QDialog { background-color: #1e1e2e; }
QToolTip {
    background-color: #313244; color: #cdd6f4;
    border: 1px solid #45475a; padding: 4px 8px; border-radius: 4px;
}

/* ── Progress bar ── */
QProgressBar {
    background-color: #313244;
    border-radius: 3px;
    border: none;
    height: 6px;
}
QProgressBar::chunk {
    background-color: #89b4fa;
    border-radius: 3px;
}
"""


# ── Profile helpers ────────────────────────────────────────────────────────────

def load_profiles() -> list:
    if os.path.exists(PROFILES_FILE):
        try:
            with open(PROFILES_FILE, "r") as f:
                return json.load(f)
        except Exception:
            return []
    return []


def save_profiles(profiles: list) -> None:
    directory = os.path.dirname(PROFILES_FILE)
    os.makedirs(directory, exist_ok=True)
    tmp = PROFILES_FILE + ".tmp"
    with open(tmp, "w") as f:
        json.dump(profiles, f, indent=2)
    os.chmod(tmp, 0o600)
    os.replace(tmp, PROFILES_FILE)


# ── Background connection worker ───────────────────────────────────────────────

class ConnectWorker(QThread):
    success = pyqtSignal(object, dict)
    failure = pyqtSignal(str)

    def __init__(self, params: dict):
        super().__init__()
        self.params = params

    def run(self):
        try:
            conn = psycopg2.connect(
                host=self.params["host"],
                port=int(self.params["port"]),
                dbname=self.params["database"],
                user=self.params["user"],
                password=self.params["password"],
                connect_timeout=10,
                application_name="pgbrowser",
            )
            self.success.emit(conn, self.params)
        except Exception as exc:
            self.failure.emit(str(exc))


# ── Login screen ───────────────────────────────────────────────────────────────

class LoginWidget(QWidget):
    connect_requested = pyqtSignal(object, dict)

    def __init__(self):
        super().__init__()
        self.profiles: list = load_profiles()
        self._worker: ConnectWorker | None = None
        self._build_ui()

    # ── UI construction ────────────────────────────────────────────────────────

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setAlignment(Qt.AlignCenter)
        root.setContentsMargins(0, 0, 0, 0)

        card = QFrame(objectName="card")
        card.setFixedWidth(480)
        cv = QVBoxLayout(card)
        cv.setContentsMargins(36, 32, 36, 32)
        cv.setSpacing(14)

        # Header
        title_lbl = QLabel("🐘  PG Browser")
        title_lbl.setFont(QFont("Segoe UI", 20, QFont.Bold))
        title_lbl.setAlignment(Qt.AlignCenter)
        title_lbl.setStyleSheet("color: #89b4fa; background: transparent;")
        cv.addWidget(title_lbl)

        sub_lbl = QLabel("PostgreSQL Schema Explorer")
        sub_lbl.setAlignment(Qt.AlignCenter)
        sub_lbl.setStyleSheet("color: #a6adc8; font-size: 12px; background: transparent;")
        cv.addWidget(sub_lbl)

        cv.addSpacing(6)

        # Saved connections row
        pr_row = QHBoxLayout()
        pr_lbl = QLabel("Saved:")
        pr_lbl.setStyleSheet("background:transparent;")
        pr_row.addWidget(pr_lbl)
        self.profile_combo = QComboBox()
        self.profile_combo.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.profile_combo.addItem("— New connection —")
        for p in self.profiles:
            self.profile_combo.addItem(self._profile_label(p))
        self.profile_combo.currentIndexChanged.connect(self._on_profile_changed)
        pr_row.addWidget(self.profile_combo)
        cv.addLayout(pr_row)

        # Form fields
        form = QFormLayout()
        form.setLabelAlignment(Qt.AlignRight)
        form.setHorizontalSpacing(12)
        form.setVerticalSpacing(10)

        self.host_edit = QLineEdit("localhost")
        self.port_edit = QLineEdit("5432")
        self.db_edit   = QLineEdit(); self.db_edit.setPlaceholderText("my_database")
        self.user_edit = QLineEdit(); self.user_edit.setPlaceholderText("postgres")
        self.pass_edit = QLineEdit(); self.pass_edit.setEchoMode(QLineEdit.Password)
        self.pass_edit.setPlaceholderText("••••••••")
        self.label_edit = QLineEdit(); self.label_edit.setPlaceholderText("Optional name for this connection")

        form.addRow("Host:",     self.host_edit)
        form.addRow("Port:",     self.port_edit)
        form.addRow("Database:", self.db_edit)
        form.addRow("User:",     self.user_edit)
        form.addRow("Password:", self.pass_edit)
        form.addRow("Label:",    self.label_edit)
        cv.addLayout(form)

        self.save_check = QCheckBox("Remember this connection")
        cv.addWidget(self.save_check)

        cv.addSpacing(4)

        # Action buttons
        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)
        self.del_btn = QPushButton("Delete", objectName="danger")
        self.del_btn.setToolTip("Remove saved connection")
        self.del_btn.setVisible(False)
        self.del_btn.clicked.connect(self._delete_profile)
        btn_row.addWidget(self.del_btn)
        btn_row.addStretch()
        self.connect_btn = QPushButton("Connect")
        self.connect_btn.setDefault(True)
        self.connect_btn.clicked.connect(self._do_connect)
        btn_row.addWidget(self.connect_btn)
        cv.addLayout(btn_row)

        self.status_lbl = QLabel("")
        self.status_lbl.setAlignment(Qt.AlignCenter)
        self.status_lbl.setWordWrap(True)
        self.status_lbl.setStyleSheet(
            "color: #f38ba8; font-size: 12px; background: transparent;"
        )
        cv.addWidget(self.status_lbl)

        root.addWidget(card, alignment=Qt.AlignCenter)
        self.pass_edit.returnPressed.connect(self._do_connect)

    # ── Helpers ────────────────────────────────────────────────────────────────

    @staticmethod
    def _profile_label(p: dict) -> str:
        return p.get("name") or f"{p['user']}@{p['host']}/{p['database']}"

    def _on_profile_changed(self, idx: int):
        if idx == 0:
            self.host_edit.setText("localhost")
            self.port_edit.setText("5432")
            self.db_edit.clear()
            self.user_edit.clear()
            self.pass_edit.clear()
            self.label_edit.clear()
            self.del_btn.setVisible(False)
        else:
            p = self.profiles[idx - 1]
            self.host_edit.setText(p.get("host", "localhost"))
            self.port_edit.setText(str(p.get("port", 5432)))
            self.db_edit.setText(p.get("database", ""))
            self.user_edit.setText(p.get("user", ""))
            self.pass_edit.setText(p.get("password", ""))
            self.label_edit.setText(p.get("name", ""))
            self.del_btn.setVisible(True)

    def _delete_profile(self):
        idx = self.profile_combo.currentIndex()
        if idx == 0:
            return
        name = self.profile_combo.currentText()
        if (
            QMessageBox.question(
                self, "Delete Profile",
                f'Remove saved connection "{name}"?',
                QMessageBox.Yes | QMessageBox.No,
            )
            == QMessageBox.Yes
        ):
            self.profiles.pop(idx - 1)
            save_profiles(self.profiles)
            self.profile_combo.removeItem(idx)
            self.profile_combo.setCurrentIndex(0)

    # ── Connect ────────────────────────────────────────────────────────────────

    def _do_connect(self):
        host = self.host_edit.text().strip()
        port = self.port_edit.text().strip()
        database = self.db_edit.text().strip()
        user = self.user_edit.text().strip()
        password = self.pass_edit.text()

        if not all([host, port, database, user]):
            self.status_lbl.setText("Host, port, database and user are required.")
            return
        try:
            int(port)
        except ValueError:
            self.status_lbl.setText("Port must be a number.")
            return

        params = {
            "host": host, "port": port,
            "database": database, "user": user, "password": password,
            "name": self.label_edit.text().strip(),
        }
        self.connect_btn.setEnabled(False)
        self.connect_btn.setText("Connecting…")
        self.status_lbl.setText("")
        self.status_lbl.setStyleSheet(
            "color: #a6adc8; font-size: 12px; background: transparent;"
        )
        self.status_lbl.setText("Connecting…")

        self._worker = ConnectWorker(params)
        self._worker.success.connect(self._on_connected)
        self._worker.failure.connect(self._on_error)
        self._worker.start()

    def _on_connected(self, conn, params):
        self.connect_btn.setEnabled(True)
        self.connect_btn.setText("Connect")
        self.status_lbl.setText("")

        if self.save_check.isChecked():
            label = params.get("name") or self._profile_label(params)
            params["name"] = label
            cur_idx = self.profile_combo.currentIndex()
            if cur_idx > 0:
                self.profiles[cur_idx - 1] = params
                self.profile_combo.setItemText(cur_idx, label)
            else:
                self.profiles.append(params)
                self.profile_combo.addItem(label)
            save_profiles(self.profiles)

        self.connect_requested.emit(conn, params)

    def _on_error(self, err: str):
        self.connect_btn.setEnabled(True)
        self.connect_btn.setText("Connect")
        self.status_lbl.setStyleSheet(
            "color: #f38ba8; font-size: 12px; background: transparent;"
        )
        self.status_lbl.setText(f"❌  {err}")


# ── Data Dictionary export worker ─────────────────────────────────────────────

class DictExportWorker(QThread):
    progress = pyqtSignal(int, int, str)   # current, total, table_name
    finished = pyqtSignal(str)             # output path
    error    = pyqtSignal(str)

    def __init__(self, params: dict, schema: str, tables: list, path: str):
        super().__init__()
        self.params = params
        self.schema = schema
        self.tables = tables
        self.path   = path

    def run(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            self.error.emit(
                "openpyxl is not installed.\nRun:  pip install openpyxl"
            )
            return

        try:
            conn = psycopg2.connect(
                host=self.params["host"],
                port=int(self.params["port"]),
                dbname=self.params["database"],
                user=self.params["user"],
                password=self.params["password"],
                connect_timeout=10,
                application_name="pgbrowser_export",
            )
        except Exception as exc:
            self.error.emit(f"Connection failed: {exc}")
            return

        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)   # drop default empty sheet

            hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
            hdr_fill  = PatternFill("solid", fgColor="1F3864")
            hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin      = Side(style="thin", color="D9D9D9")
            brd       = Border(left=thin, right=thin, top=thin, bottom=thin)
            pk_fill   = PatternFill("solid", fgColor="FFF2CC")
            fk_fill   = PatternFill("solid", fgColor="DEEAF1")

            HEADERS    = [
                "#", "Column Name", "Data Type", "UDT Name",
                "Max Length", "Precision", "Scale",
                "Nullable", "Default Value", "Constraints", "Description",
            ]
            COL_WIDTHS = [5, 25, 18, 18, 12, 12, 8, 10, 30, 18, 40]

            for idx, table in enumerate(self.tables, 1):
                self.progress.emit(idx, len(self.tables), table)

                cur = conn.cursor(cursor_factory=RealDictCursor)
                cur.execute("""
                    SELECT
                        c.ordinal_position,
                        c.column_name,
                        c.data_type,
                        c.udt_name,
                        c.character_maximum_length,
                        c.numeric_precision,
                        c.numeric_scale,
                        c.is_nullable,
                        c.column_default,
                        COALESCE(STRING_AGG(DISTINCT
                            CASE tc.constraint_type
                                WHEN 'PRIMARY KEY' THEN 'PK'
                                WHEN 'FOREIGN KEY' THEN 'FK'
                                WHEN 'UNIQUE'      THEN 'UQ'
                                ELSE NULL
                            END, ', '), '') AS constraints,
                        d.description
                    FROM information_schema.columns c
                    LEFT JOIN information_schema.key_column_usage kcu
                        ON  c.table_schema = kcu.table_schema
                        AND c.table_name   = kcu.table_name
                        AND c.column_name  = kcu.column_name
                    LEFT JOIN information_schema.table_constraints tc
                        ON  kcu.constraint_name   = tc.constraint_name
                        AND kcu.constraint_schema = tc.constraint_schema
                        AND kcu.table_name        = tc.table_name
                    LEFT JOIN pg_catalog.pg_description d
                        ON  d.objoid = (
                                SELECT oid FROM pg_class
                                JOIN   pg_namespace
                                       ON pg_namespace.oid = pg_class.relnamespace
                                WHERE  pg_class.relname   = %s
                                  AND  pg_namespace.nspname = %s
                            )
                        AND d.objsubid = c.ordinal_position
                    WHERE c.table_schema = %s AND c.table_name = %s
                    GROUP BY c.ordinal_position, c.column_name, c.data_type,
                             c.udt_name, c.character_maximum_length,
                             c.numeric_precision, c.numeric_scale,
                             c.is_nullable, c.column_default, d.description
                    ORDER BY c.ordinal_position
                """, (table, self.schema, self.schema, table))
                rows = cur.fetchall()

                # Sheet name: max 31 chars, must be unique
                sheet_name  = table[:31]
                used_names  = {ws.title for ws in wb.worksheets}
                if sheet_name in used_names:
                    sheet_name = f"{table[:28]}_{idx}"

                ws = wb.create_sheet(title=sheet_name)

                # Title row
                ws.merge_cells("A1:K1")
                tc           = ws["A1"]
                tc.value     = f"Data Dictionary — {self.schema}.{table}"
                tc.font      = Font(name="Calibri", bold=True, size=13, color="1F3864")
                tc.alignment = Alignment(horizontal="left", vertical="center")
                ws.row_dimensions[1].height = 22

                # Header row
                for ci, (hdr, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
                    cell            = ws.cell(row=2, column=ci, value=hdr)
                    cell.font       = hdr_font
                    cell.fill       = hdr_fill
                    cell.alignment  = hdr_align
                    cell.border     = brd
                    ws.column_dimensions[cell.column_letter].width = w
                ws.row_dimensions[2].height = 18

                # Data rows
                for ri, row in enumerate(rows, 3):
                    constraints = row["constraints"] or ""
                    vals = [
                        row["ordinal_position"],
                        row["column_name"],
                        row["data_type"],
                        row["udt_name"],
                        row["character_maximum_length"],
                        row["numeric_precision"],
                        row["numeric_scale"],
                        row["is_nullable"],
                        row["column_default"] or "",
                        constraints,
                        row["description"] or "",
                    ]
                    for ci, val in enumerate(vals, 1):
                        cell           = ws.cell(row=ri, column=ci, value=val)
                        cell.border    = brd
                        cell.alignment = Alignment(
                            vertical="center",
                            wrap_text=(ci in (9, 11)),
                        )
                        if ci == 10:
                            if "PK" in constraints:
                                cell.fill = pk_fill
                            elif "FK" in constraints:
                                cell.fill = fk_fill
                    ws.row_dimensions[ri].height = 15

                ws.freeze_panes = "A3"

            wb.save(self.path)
            self.finished.emit(self.path)

        except Exception as exc:
            self.error.emit(str(exc))
        finally:
            try:
                conn.close()
            except Exception:
                pass


# ── Main browser window ────────────────────────────────────────────────────────

class BrowserWindow(QWidget):
    logout_requested = pyqtSignal()

    def __init__(self, conn, params: dict):
        super().__init__()
        self.conn = conn
        self.params = params
        self._all_tables: list[tuple[str, str]] = []
        self._dd_all_tables: list[tuple[str, str]] = []
        self._dd_export_worker: DictExportWorker | None = None
        self._build_ui()
        self._load_schemas()

    # ── UI construction ────────────────────────────────────────────────────────

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Top bar
        topbar = QFrame()
        topbar.setFixedHeight(54)
        topbar.setStyleSheet(
            "QFrame { background-color: #181825; border-bottom: 1px solid #313244; }"
        )
        tb = QHBoxLayout(topbar)
        tb.setContentsMargins(16, 0, 16, 0)
        tb.setSpacing(10)

        logo = QLabel("🐘  PG Browser")
        logo.setFont(QFont("Segoe UI", 13, QFont.Bold))
        logo.setStyleSheet("color: #89b4fa; background: transparent;")
        tb.addWidget(logo)

        sep = QFrame(); sep.setFrameShape(QFrame.VLine)
        sep.setStyleSheet("QFrame { color: #45475a; background: #45475a; }")
        sep.setFixedWidth(1)
        tb.addWidget(sep)

        conn_info = QLabel(
            f"{self.params['user']}@{self.params['host']}:{self.params['port']}"
            f"  /  {self.params['database']}"
        )
        conn_info.setStyleSheet("color: #a6adc8; font-size: 12px; background: transparent;")
        tb.addWidget(conn_info)

        tb.addStretch()

        schema_lbl = QLabel("Schema:")
        schema_lbl.setStyleSheet("color: #a6adc8; font-size: 12px; background: transparent;")
        tb.addWidget(schema_lbl)

        self.schema_combo = QComboBox()
        self.schema_combo.setMinimumWidth(160)
        self.schema_combo.setFixedHeight(32)
        self.schema_combo.currentTextChanged.connect(self._on_schema_changed)
        tb.addWidget(self.schema_combo)

        sep2 = QFrame(); sep2.setFrameShape(QFrame.VLine)
        sep2.setStyleSheet("QFrame { color: #45475a; background: #45475a; }")
        sep2.setFixedWidth(1)
        tb.addWidget(sep2)

        disc_btn = QPushButton("⏏  Disconnect", objectName="secondary")
        disc_btn.setFixedHeight(32)
        disc_btn.clicked.connect(self._disconnect)
        tb.addWidget(disc_btn)

        root.addWidget(topbar)

        # ── Content splitter
        splitter = QSplitter(Qt.Horizontal)
        splitter.setChildrenCollapsible(False)
        splitter.setHandleWidth(1)

        # ── Left panel: table list
        left = QWidget()
        left.setMinimumWidth(180)
        left.setMaximumWidth(340)
        lv = QVBoxLayout(left)
        lv.setContentsMargins(8, 10, 4, 8)
        lv.setSpacing(6)

        tbl_title = QLabel("Tables & Views")
        tbl_title.setFont(QFont("Segoe UI", 11, QFont.Bold))
        tbl_title.setStyleSheet("color: #a6adc8; background: transparent; padding: 2px 4px;")
        lv.addWidget(tbl_title)

        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("🔍  Filter…")
        self.search_edit.textChanged.connect(self._filter_tables)
        lv.addWidget(self.search_edit)

        self.table_list = QListWidget()
        self.table_list.currentItemChanged.connect(self._on_table_selected)
        lv.addWidget(self.table_list)

        self.count_lbl = QLabel("")
        self.count_lbl.setStyleSheet(
            "color: #585b70; font-size: 11px; background: transparent;"
        )
        lv.addWidget(self.count_lbl)

        splitter.addWidget(left)

        # ── Right panel: table details
        right = QWidget()
        rv = QVBoxLayout(right)
        rv.setContentsMargins(4, 10, 10, 10)
        rv.setSpacing(6)

        self.detail_title = QLabel("Select a table to inspect")
        self.detail_title.setFont(QFont("Segoe UI", 15, QFont.Bold))
        self.detail_title.setStyleSheet(
            "color: #585b70; background: transparent; padding: 2px 4px;"
        )
        rv.addWidget(self.detail_title)

        self.detail_meta = QLabel("")
        self.detail_meta.setStyleSheet(
            "color: #6c7086; font-size: 12px; background: transparent; padding: 0 4px 4px;"
        )
        rv.addWidget(self.detail_meta)

        # Tabs
        self.tabs = QTabWidget()
        self.tabs.setDocumentMode(True)

        # Columns tab
        self.col_table = self._make_table(
            ["#", "Column", "Type", "Nullable", "Default", "Constraints"]
        )
        self.tabs.addTab(self._wrap(self.col_table), "Columns")

        # Indexes tab
        self.idx_table = self._make_table(
            ["Index Name", "Method", "Unique", "Primary", "Columns"]
        )
        self.tabs.addTab(self._wrap(self.idx_table), "Indexes")

        # Foreign keys tab
        self.fk_table = self._make_table(
            ["Constraint", "Column(s)", "References", "On Update", "On Delete"]
        )
        self.tabs.addTab(self._wrap(self.fk_table), "Foreign Keys")

        rv.addWidget(self.tabs)
        splitter.addWidget(right)
        splitter.setSizes([240, 960])

        browse_tab = QWidget()
        blt = QVBoxLayout(browse_tab)
        blt.setContentsMargins(0, 0, 0, 0)
        blt.setSpacing(0)
        blt.addWidget(splitter)

        self._main_tabs = QTabWidget()
        self._main_tabs.setDocumentMode(True)
        self._main_tabs.addTab(browse_tab, "Browse Tables")
        self._main_tabs.addTab(self._build_dict_tab(), "Data Dictionary")
        root.addWidget(self._main_tabs)

    @staticmethod
    def _make_table(headers: list[str]) -> QTableWidget:
        t = QTableWidget()
        t.setColumnCount(len(headers))
        t.setHorizontalHeaderLabels(headers)
        t.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        t.horizontalHeader().setStretchLastSection(True)
        t.verticalHeader().setVisible(False)
        t.setEditTriggers(QAbstractItemView.NoEditTriggers)
        t.setSelectionBehavior(QAbstractItemView.SelectRows)
        t.setAlternatingRowColors(True)
        t.setWordWrap(False)
        t.setShowGrid(False)
        return t

    @staticmethod
    def _wrap(widget: QWidget) -> QWidget:
        w = QWidget()
        v = QVBoxLayout(w)
        v.setContentsMargins(0, 8, 0, 0)
        v.addWidget(widget)
        return w

    # ── DB helpers ─────────────────────────────────────────────────────────────

    def _exec(self, query: str, params: tuple = ()) -> list:
        try:
            cur = self.conn.cursor(cursor_factory=RealDictCursor)
            cur.execute(query, params)
            return cur.fetchall()
        except Exception as exc:
            self.conn.rollback()
            raise exc

    # ── Schema loading ─────────────────────────────────────────────────────────

    def _load_schemas(self):
        rows = self._exec("""
            SELECT schema_name
            FROM   information_schema.schemata
            WHERE  schema_name NOT IN ('pg_catalog', 'information_schema', 'pg_toast')
              AND  schema_name NOT LIKE 'pg_temp_%%'
              AND  schema_name NOT LIKE 'pg_toast_temp_%%'
            ORDER BY
                CASE WHEN schema_name = 'public' THEN 0 ELSE 1 END,
                schema_name
        """)
        self.schema_combo.blockSignals(True)
        self.schema_combo.clear()
        for row in rows:
            self.schema_combo.addItem(row["schema_name"])
        self.schema_combo.blockSignals(False)

        idx = self.schema_combo.findText("public")
        start = idx if idx >= 0 else 0
        self.schema_combo.setCurrentIndex(start)
        self._load_tables(self.schema_combo.currentText())

    def _load_tables(self, schema: str):
        rows = self._exec("""
            SELECT table_name, table_type
            FROM   information_schema.tables
            WHERE  table_schema = %s
            ORDER BY table_type DESC, table_name
        """, (schema,))
        self._all_tables = [(r["table_name"], r["table_type"]) for r in rows]
        self._render_tables(self._all_tables)

    def _render_tables(self, tables: list[tuple[str, str]]):
        self.table_list.clear()
        for name, ttype in tables:
            icon = "▦ " if ttype == "BASE TABLE" else "◫ "
            item = QListWidgetItem(icon + name)
            item.setData(Qt.UserRole,     ttype)
            item.setData(Qt.UserRole + 1, name)
            tip = "TABLE" if ttype == "BASE TABLE" else "VIEW"
            item.setToolTip(tip)
            self.table_list.addItem(item)
        n = len(tables)
        self.count_lbl.setText(f"{n} object{'s' if n != 1 else ''}")

    def _filter_tables(self, text: str):
        lo = text.lower()
        filtered = [(n, t) for n, t in self._all_tables if lo in n.lower()]
        self._render_tables(filtered)

    def _on_schema_changed(self, schema: str):
        if not schema:
            return
        self.search_edit.clear()
        self._load_tables(schema)
        self.detail_title.setText("Select a table to inspect")
        self.detail_title.setStyleSheet(
            "color: #585b70; background: transparent; padding: 2px 4px;"
        )
        self.detail_meta.setText("")
        for t in (self.col_table, self.idx_table, self.fk_table):
            t.setRowCount(0)

    # ── Table selection ────────────────────────────────────────────────────────

    def _on_table_selected(self, current, _previous):
        if current is None:
            return
        name  = current.data(Qt.UserRole + 1)
        ttype = current.data(Qt.UserRole)
        if name:
            self._load_table_details(self.schema_combo.currentText(), name, ttype)

    def _load_table_details(self, schema: str, table: str, ttype: str):
        self.detail_title.setText(f"  {schema}.{table}")
        self.detail_title.setStyleSheet(
            "color: #cdd6f4; background: transparent; padding: 2px 4px;"
        )

        kind = "TABLE" if ttype == "BASE TABLE" else "VIEW"
        if ttype == "BASE TABLE":
            try:
                rows = self._exec("""
                    SELECT reltuples::bigint AS estimate
                    FROM   pg_class c
                    JOIN   pg_namespace n ON n.oid = c.relnamespace
                    WHERE  n.nspname = %s AND c.relname = %s
                """, (schema, table))
                est = rows[0]["estimate"] if rows else 0
                meta = f"{kind}   •   ~{est:,} rows (estimated)"
            except Exception:
                meta = kind
        else:
            meta = kind
        self.detail_meta.setText(meta)

        self._load_columns(schema, table)
        self._load_indexes(schema, table)
        self._load_foreign_keys(schema, table)

    # ── Columns ────────────────────────────────────────────────────────────────

    def _load_columns(self, schema: str, table: str):
        rows = self._exec("""
            SELECT
                c.ordinal_position,
                c.column_name,
                CASE
                    WHEN c.character_maximum_length IS NOT NULL
                        THEN c.data_type || '(' || c.character_maximum_length || ')'
                    WHEN c.numeric_precision IS NOT NULL
                         AND c.data_type IN ('numeric', 'decimal')
                        THEN c.data_type
                             || '(' || c.numeric_precision
                             || ',' || COALESCE(c.numeric_scale::text, '0') || ')'
                    ELSE c.data_type
                END AS display_type,
                c.is_nullable,
                c.column_default,
                COALESCE(STRING_AGG(DISTINCT
                    CASE tc.constraint_type
                        WHEN 'PRIMARY KEY' THEN 'PK'
                        WHEN 'FOREIGN KEY' THEN 'FK'
                        WHEN 'UNIQUE'      THEN 'UQ'
                        ELSE tc.constraint_type
                    END,
                    ', '), '') AS constraints
            FROM information_schema.columns c
            LEFT JOIN information_schema.key_column_usage kcu
                ON  c.table_schema = kcu.table_schema
                AND c.table_name   = kcu.table_name
                AND c.column_name  = kcu.column_name
            LEFT JOIN information_schema.table_constraints tc
                ON  kcu.constraint_name   = tc.constraint_name
                AND kcu.constraint_schema = tc.constraint_schema
                AND kcu.table_name        = tc.table_name
            WHERE c.table_schema = %s AND c.table_name = %s
            GROUP BY c.ordinal_position, c.column_name, c.data_type,
                     c.character_maximum_length, c.numeric_precision,
                     c.numeric_scale, c.is_nullable, c.column_default
            ORDER BY c.ordinal_position
        """, (schema, table))

        t = self.col_table
        t.setRowCount(len(rows))
        for r, row in enumerate(rows):
            nullable    = row["is_nullable"] == "YES"
            default_val = row["column_default"] or ""
            constraints = row["constraints"] or ""

            vals = [
                str(row["ordinal_position"]),
                row["column_name"],
                row["display_type"],
                "YES" if nullable else "NO",
                default_val,
                constraints,
            ]
            for c, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                # Colour coding
                if c == 3:  # Nullable
                    item.setForeground(
                        QColor("#a6e3a1") if nullable else QColor("#f38ba8")
                    )
                elif c == 5 and "PK" in constraints:
                    item.setForeground(QColor("#f9e2af"))
                    item.setFont(QFont("", -1, QFont.Bold))
                elif c == 5 and "FK" in constraints:
                    item.setForeground(QColor("#89dceb"))
                t.setItem(r, c, item)

        t.resizeColumnsToContents()
        t.setColumnWidth(0, 40)   # # column
        if t.columnWidth(1) < 150:
            t.setColumnWidth(1, 150)
        if t.columnWidth(2) < 130:
            t.setColumnWidth(2, 130)

    # ── Indexes ────────────────────────────────────────────────────────────────

    def _load_indexes(self, schema: str, table: str):
        rows = self._exec("""
            SELECT
                pidx.indexname   AS index_name,
                UPPER(COALESCE(
                    SUBSTRING(pidx.indexdef FROM ' USING (\\w+)'),
                    'btree'
                ))               AS index_method,
                ix.indisunique   AS is_unique,
                ix.indisprimary  AS is_primary,
                ARRAY_TO_STRING(ARRAY(
                    SELECT a.attname
                    FROM   pg_attribute a
                    WHERE  a.attrelid = ix.indrelid
                      AND  a.attnum   = ANY(ix.indkey)
                      AND  a.attnum  > 0
                    ORDER BY array_position(ix.indkey, a.attnum::smallint)
                ), ', ') AS columns
            FROM   pg_indexes  pidx
            JOIN   pg_class    ic ON ic.relname = pidx.indexname
            JOIN   pg_namespace n  ON n.oid = ic.relnamespace
                                  AND n.nspname = pidx.schemaname
            JOIN   pg_index    ix  ON ix.indexrelid = ic.oid
            WHERE  pidx.schemaname = %s AND pidx.tablename = %s
            ORDER BY ix.indisprimary DESC, pidx.indexname
        """, (schema, table))

        t = self.idx_table
        t.setRowCount(len(rows))
        for r, row in enumerate(rows):
            vals = [
                row["index_name"],
                row["index_method"].upper(),
                "YES" if row["is_unique"]   else "NO",
                "YES" if row["is_primary"]  else "NO",
                row["columns"],
            ]
            for c, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                if c == 2:
                    item.setForeground(
                        QColor("#a6e3a1") if row["is_unique"] else QColor("#6c7086")
                    )
                if c == 3 and row["is_primary"]:
                    item.setForeground(QColor("#f9e2af"))
                    item.setFont(QFont("", -1, QFont.Bold))
                t.setItem(r, c, item)
        t.resizeColumnsToContents()

    # ── Foreign keys ───────────────────────────────────────────────────────────

    def _load_foreign_keys(self, schema: str, table: str):
        rows = self._exec("""
            SELECT
                tc.constraint_name,
                STRING_AGG(kcu.column_name, ', '
                           ORDER BY kcu.ordinal_position) AS columns,
                ccu.table_schema || '.' || ccu.table_name AS ref_table,
                STRING_AGG(ccu.column_name, ', '
                           ORDER BY kcu.ordinal_position) AS ref_columns,
                rc.update_rule,
                rc.delete_rule
            FROM information_schema.table_constraints tc
            JOIN information_schema.key_column_usage kcu
                ON  tc.constraint_name   = kcu.constraint_name
                AND tc.constraint_schema = kcu.constraint_schema
                AND tc.table_name        = kcu.table_name
            JOIN information_schema.referential_constraints rc
                ON  tc.constraint_name   = rc.constraint_name
                AND tc.constraint_schema = rc.constraint_schema
            JOIN information_schema.constraint_column_usage ccu
                ON  rc.unique_constraint_name   = ccu.constraint_name
                AND rc.unique_constraint_schema = ccu.constraint_schema
            WHERE tc.constraint_type = 'FOREIGN KEY'
              AND tc.table_schema = %s
              AND tc.table_name   = %s
            GROUP BY tc.constraint_name, ccu.table_schema, ccu.table_name,
                     rc.update_rule, rc.delete_rule
            ORDER BY tc.constraint_name
        """, (schema, table))

        t = self.fk_table
        t.setRowCount(len(rows))
        for r, row in enumerate(rows):
            ref = f"{row['ref_table']} ({row['ref_columns']})"
            vals = [
                row["constraint_name"],
                row["columns"],
                ref,
                row["update_rule"],
                row["delete_rule"],
            ]
            for c, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                if c == 2:
                    item.setForeground(QColor("#89dceb"))
                t.setItem(r, c, item)
        t.resizeColumnsToContents()

    # ── Data Dictionary tab ────────────────────────────────────────────────────

    def _build_dict_tab(self) -> QWidget:
        w = QWidget()
        v = QVBoxLayout(w)
        v.setContentsMargins(16, 14, 16, 14)
        v.setSpacing(8)

        # Title
        title = QLabel("Data Dictionary Generator")
        title.setFont(QFont("Segoe UI", 14, QFont.Bold))
        title.setStyleSheet("color: #cdd6f4; background: transparent;")
        v.addWidget(title)

        sub = QLabel(
            "Load tables, select the ones to include, then generate an Excel file "
            "with a dedicated sheet per table."
        )
        sub.setWordWrap(True)
        sub.setStyleSheet("color: #a6adc8; font-size: 12px; background: transparent;")
        v.addWidget(sub)

        sep = QFrame(); sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet("QFrame { color: #313244; background: #313244; }")
        sep.setFixedHeight(1)
        v.addWidget(sep)

        # Search + Load row
        sr = QHBoxLayout()
        sr.setSpacing(8)
        self.dd_search = QLineEdit()
        self.dd_search.setPlaceholderText("🔍  Search tables…")
        self.dd_search.textChanged.connect(self._dd_filter_tables)
        sr.addWidget(self.dd_search)

        self.dd_load_btn = QPushButton("Load Tables")
        self.dd_load_btn.setFixedWidth(120)
        self.dd_load_btn.clicked.connect(self._dd_load_tables)
        sr.addWidget(self.dd_load_btn)
        v.addLayout(sr)

        # Select All / Deselect All + count
        sel_row = QHBoxLayout()
        sel_row.setSpacing(6)

        sel_all_btn = QPushButton("Select All", objectName="secondary")
        sel_all_btn.setFixedWidth(100)
        sel_all_btn.clicked.connect(self._dd_select_all)
        sel_row.addWidget(sel_all_btn)

        desel_all_btn = QPushButton("Deselect All", objectName="secondary")
        desel_all_btn.setFixedWidth(110)
        desel_all_btn.clicked.connect(self._dd_deselect_all)
        sel_row.addWidget(desel_all_btn)

        sel_row.addStretch()

        self.dd_sel_count = QLabel("0 selected")
        self.dd_sel_count.setStyleSheet(
            "color: #a6adc8; font-size: 12px; background: transparent;"
        )
        sel_row.addWidget(self.dd_sel_count)
        v.addLayout(sel_row)

        # Table list with checkboxes
        self.dd_list = QListWidget()
        self.dd_list.setAlternatingRowColors(True)
        self.dd_list.itemChanged.connect(self._dd_on_item_changed)
        v.addWidget(self.dd_list)

        # Generate button row
        gen_row = QHBoxLayout()
        gen_row.setSpacing(10)

        self.dd_gen_btn = QPushButton("⬇  Generate Excel")
        self.dd_gen_btn.setFixedWidth(180)
        self.dd_gen_btn.clicked.connect(self._dd_generate)
        gen_row.addWidget(self.dd_gen_btn)

        gen_row.addStretch()
        v.addLayout(gen_row)

        # Progress bar
        self.dd_progress = QProgressBar()
        self.dd_progress.setVisible(False)
        self.dd_progress.setFixedHeight(6)
        self.dd_progress.setTextVisible(False)
        v.addWidget(self.dd_progress)

        # Status label
        self.dd_status = QLabel("")
        self.dd_status.setWordWrap(True)
        self.dd_status.setStyleSheet(
            "color: #a6adc8; font-size: 12px; background: transparent;"
        )
        v.addWidget(self.dd_status)

        return w

    # ── DD helpers ─────────────────────────────────────────────────────────────

    def _dd_load_tables(self):
        schema = self.schema_combo.currentText()
        if not schema:
            return
        self.dd_load_btn.setEnabled(False)
        self.dd_status.setStyleSheet(
            "color: #a6adc8; font-size: 12px; background: transparent;"
        )
        try:
            rows = self._exec("""
                SELECT table_name, table_type
                FROM   information_schema.tables
                WHERE  table_schema = %s
                ORDER BY table_type DESC, table_name
            """, (schema,))
            self._dd_all_tables = [(r["table_name"], r["table_type"]) for r in rows]
        except Exception as exc:
            self.dd_status.setStyleSheet(
                "color: #f38ba8; font-size: 12px; background: transparent;"
            )
            self.dd_status.setText(f"❌  Error loading tables: {exc}")
            self.dd_load_btn.setEnabled(True)
            return
        self.dd_load_btn.setEnabled(True)
        self.dd_search.clear()
        self._dd_render_tables(self._dd_all_tables)
        n = len(self._dd_all_tables)
        self.dd_status.setText(
            f"Loaded {n} object{'s' if n != 1 else ''} from schema '{schema}'"
        )

    def _dd_render_tables(self, tables: list):
        self.dd_list.blockSignals(True)
        self.dd_list.clear()
        for name, ttype in tables:
            icon = "▦ " if ttype == "BASE TABLE" else "◫ "
            item = QListWidgetItem(icon + name)
            item.setData(Qt.UserRole, name)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Unchecked)
            tip = "TABLE" if ttype == "BASE TABLE" else "VIEW"
            item.setToolTip(tip)
            self.dd_list.addItem(item)
        self.dd_list.blockSignals(False)
        self._dd_update_count()

    def _dd_filter_tables(self, text: str):
        lo = text.lower()
        filtered = [(n, t) for n, t in self._dd_all_tables if lo in n.lower()]
        self._dd_render_tables(filtered)

    def _dd_select_all(self):
        self.dd_list.blockSignals(True)
        for i in range(self.dd_list.count()):
            self.dd_list.item(i).setCheckState(Qt.Checked)
        self.dd_list.blockSignals(False)
        self._dd_update_count()

    def _dd_deselect_all(self):
        self.dd_list.blockSignals(True)
        for i in range(self.dd_list.count()):
            self.dd_list.item(i).setCheckState(Qt.Unchecked)
        self.dd_list.blockSignals(False)
        self._dd_update_count()

    def _dd_on_item_changed(self, _item):
        self._dd_update_count()

    def _dd_update_count(self):
        n = sum(
            1 for i in range(self.dd_list.count())
            if self.dd_list.item(i).checkState() == Qt.Checked
        )
        self.dd_sel_count.setText(f"{n} selected")

    def _dd_generate(self):
        selected = [
            self.dd_list.item(i).data(Qt.UserRole)
            for i in range(self.dd_list.count())
            if self.dd_list.item(i).checkState() == Qt.Checked
        ]
        if not selected:
            QMessageBox.warning(
                self, "No Tables Selected",
                "Please select at least one table before generating."
            )
            return

        schema = self.schema_combo.currentText()
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Data Dictionary",
            f"data_dictionary_{schema}.xlsx",
            "Excel Files (*.xlsx)",
        )
        if not path:
            return
        if not path.endswith(".xlsx"):
            path += ".xlsx"

        self.dd_gen_btn.setEnabled(False)
        self.dd_progress.setRange(0, len(selected))
        self.dd_progress.setValue(0)
        self.dd_progress.setVisible(True)
        self.dd_status.setStyleSheet(
            "color: #a6adc8; font-size: 12px; background: transparent;"
        )
        self.dd_status.setText(f"Generating data dictionary for {len(selected)} table(s)…")

        self._dd_export_worker = DictExportWorker(self.params, schema, selected, path)
        self._dd_export_worker.progress.connect(self._dd_on_progress)
        self._dd_export_worker.finished.connect(self._dd_on_finished)
        self._dd_export_worker.error.connect(self._dd_on_error)
        self._dd_export_worker.start()

    def _dd_on_progress(self, current: int, total: int, name: str):
        self.dd_progress.setValue(current)
        self.dd_status.setText(f"Processing {current}/{total}: {name}")

    def _dd_on_finished(self, path: str):
        self.dd_gen_btn.setEnabled(True)
        self.dd_progress.setVisible(False)
        self.dd_status.setStyleSheet(
            "color: #a6e3a1; font-size: 12px; background: transparent;"
        )
        self.dd_status.setText(f"✓  Saved: {path}")

    def _dd_on_error(self, err: str):
        self.dd_gen_btn.setEnabled(True)
        self.dd_progress.setVisible(False)
        self.dd_status.setStyleSheet(
            "color: #f38ba8; font-size: 12px; background: transparent;"
        )
        self.dd_status.setText(f"❌  {err}")

    # ── Disconnect ─────────────────────────────────────────────────────────────

    def _disconnect(self):
        try:
            self.conn.close()
        except Exception:
            pass
        self.logout_requested.emit()


# ── Application shell ──────────────────────────────────────────────────────────

class AppShell(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("PG Browser")
        self.resize(1280, 760)
        self.setMinimumSize(900, 550)

        self._stack = QStackedWidget()
        self.setCentralWidget(self._stack)

        self._login = LoginWidget()
        self._login.connect_requested.connect(self._on_connected)
        self._stack.addWidget(self._login)
        self._browser: BrowserWindow | None = None

        if not HAS_PSYCOPG2:
            QMessageBox.critical(
                self,
                "Missing dependency",
                "psycopg2 is not installed.\n\n"
                "Run:  pip install psycopg2-binary",
            )

    def _on_connected(self, conn, params: dict):
        if self._browser is not None:
            self._stack.removeWidget(self._browser)
            self._browser.deleteLater()

        self._browser = BrowserWindow(conn, params)
        self._browser.logout_requested.connect(self._on_logout)
        self._stack.addWidget(self._browser)
        self._stack.setCurrentWidget(self._browser)
        self.setWindowTitle(
            f"PG Browser — {params['user']}@{params['host']}/{params['database']}"
        )

    def _on_logout(self):
        self._stack.setCurrentWidget(self._login)
        self.setWindowTitle("PG Browser")


# ── Entry point ────────────────────────────────────────────────────────────────

def main():
    app = QApplication(sys.argv)
    app.setApplicationName("PG Browser")
    app.setApplicationDisplayName("PG Browser")
    app.setStyle("Fusion")
    app.setStyleSheet(STYLESHEET)

    window = AppShell()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
